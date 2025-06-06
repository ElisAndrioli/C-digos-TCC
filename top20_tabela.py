import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Cor de fundo verde e texto branco em negrito (como na imagem)
cabecalho_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
cabecalho_font = Font(bold=True)

# Lista arquivos CSV
arquivos = glob.glob("*_top20_kmers.csv")

for arquivo in arquivos:
    familia = arquivo.replace("_top20_kmers.csv", "")
    df = pd.read_csv(arquivo)

    # Ordena e adiciona a posição
    df = df.sort_values(by='frequencia', ascending=False).reset_index(drop=True)
    df.insert(0, "Posição", [f"{i+1}º lugar" for i in range(len(df))])

    # Salva como Excel temporariamente
    nome_excel = f"{familia}_tabelaTop20_formatada.xlsx"
    df.to_excel(nome_excel, index=False)

    # Aplica estilo com openpyxl
    wb = load_workbook(nome_excel)
    ws = wb.active

    for cell in ws[1]:  # Primeira linha (cabeçalho)
        cell.fill = cabecalho_fill
        cell.font = cabecalho_font

    wb.save(nome_excel)
    print(f"Tabela formatada salva: {nome_excel}")
